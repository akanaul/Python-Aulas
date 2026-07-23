"""
======================================
GABARITO MANUAL COMENTADO PASSO A PASSO
======================================
"""

# Exemplo de Solução:
# def main():
#     # ==============================================================================
# 🛑 REGRA DE OURO PARA O APRENDIZADO DE LOGÍSTICA:
# 1. Tente resolver cada exercício SOZINHO primeiro.
# 2. Se travar, use a dica do exercício ou peça ajuda ao Antigravity (IA copiloto)
#    usando os prompts de dica sugeridos (sem pedir a resposta direta!).
# 3. O GABARITO DETALHADO E COMENTADO está no final deste arquivo.
#    Consulte o gabarito APENAS se tiver tentado de verdade e continuar travado!
# ==============================================================================

"""
Exercícios — Aula 11: Excel com openpyxl e pandas
Curso: Python + IA para Automação em Logística

INSTRUÇÕES:
- Resolva cada exercício no espaço indicado
- Use o Antigravity para tirar dúvidas (mas tente primeiro!)
- O gabarito está comentado ao final do arquivo
"""
# 1. Use pandas para ler um dataframe simulado e mostrar a coluna de placas.
import pandas as pd
# Simulando a leitura: df = pd.DataFrame({'Placa': ['ABC', 'DEF'], 'Status': ['OK', 'Manutenção']})
#
# Seu código aqui
#

# 2. Filtre os veículos que tenham status "Em Manutenção".
#
# Seu código aqui
#

# 3. Use openpyxl para criar/abrir uma planilha e escrever seu nome na célula A1.
from openpyxl import Workbook
#
# Seu código aqui
#

# 4. Agrupe um dataframe de entregas por região (use groupby).
# df_entregas = pd.DataFrame({'Regiao': ['Sul', 'Sul', 'Norte'], 'Pacotes': [10, 20, 15]})
#
# Seu código aqui
#

# 5. Adicione um loop com openpyxl iterando ate max_row para formatar.
#
# Seu código aqui
#

"""
# GABARITO

# 1. Lendo com pandas e mostrando coluna
import pandas as pd
df = pd.DataFrame({'Placa': ['ABC-1234', 'XYZ-9876'], 'Status': ['OK', 'Em Manutenção']})
print(df['Placa'])

# 2. Filtrando
df_manutencao = df[df['Status'] == 'Em Manutenção']
print(df_manutencao)

# 3. Escrevendo com openpyxl
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1'] = 'Nome do Aluno'
wb.save('meu_nome.xlsx')

# 4. Groupby
df_entregas = pd.DataFrame({'Regiao': ['Sul', 'Sul', 'Norte'], 'Pacotes': [10, 20, 15]})
agrupado = df_entregas.groupby('Regiao').sum()
print(agrupado)

# 5. Loop de formatação com openpyxl
from openpyxl.styles import PatternFill
# (Assuma wb e ws ja inicializados)
# fill_green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
# for row in range(2, ws.max_row + 1):
#     if ws.cell(row, 2).value == 'Entregue':
#         ws.cell(row, 2).fill = fill_green
"""

# ---------------------------------------------------------------------------
# SISTEMA DE AUDITORIA COM IA
# ---------------------------------------------------------------------------
# INSTRUÇÕES:
# Peça para o Antigravity (ou outra IA) auditar a sua solução final usando o 
# "Prompt de Sistema de Auditoria" ensinado na aula.
# 
# GABARITO DE AUDITORIA (O que a IA deve encontrar/sugerir):
# - Uso de try/except para capturar falhas imprevistas.
# - Boas práticas de nomenclatura e legibilidade.
# - Se houver manipulação de UI/arquivos, verificar se há proteções (failsafe, close).
# ---------------------------------------------------------------------------

pass