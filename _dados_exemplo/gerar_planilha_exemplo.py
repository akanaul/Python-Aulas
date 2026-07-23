import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
import os

def gerar_planilha():
    csv_file = 'entregas_exemplo.csv'
    excel_file = 'entregas_exemplo.xlsx'
    
    # Check if CSV exists, if not relative path, maybe we run from same folder
    if not os.path.exists(csv_file):
        # Fallback to absolute path based on expected location
        csv_file = r'c:\Users\Lenovo\Documents\Curso Python e IA para automação básica\_dados_exemplo\entregas_exemplo.csv'
        excel_file = r'c:\Users\Lenovo\Documents\Curso Python e IA para automação básica\_dados_exemplo\entregas_exemplo.xlsx'

    if not os.path.exists(csv_file):
        print(f"Erro: Arquivo {csv_file} não encontrado.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Entregas"

    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)

    # Estilos
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Formatando cabeçalho
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    # Formatando o resto
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # Ajustar largura das colunas
    col_widths = {'A': 10, 'B': 20, 'C': 15, 'D': 15, 'E': 15, 'F': 15, 'G': 10, 'H': 15}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    wb.save(excel_file)
    print(f"Planilha {excel_file} gerada com sucesso!")

if __name__ == "__main__":
    gerar_planilha()
